#!/usr/bin/env python3
"""
IP Business-Development Prospect Builder
=========================================

Pulls company names from up to four public Canadian/international IP and
markets data sources, filters them to sectors of interest, and consolidates
everything into one deduplicated spreadsheet.

SOURCES & HOW THEY'RE ACCESSED
-------------------------------
1. CIPO IP Horizons bulk PATENT data   -> fully automatable (free CSV/XML download)
   https://ised-isde.canada.ca/site/canadian-intellectual-property-office/en/
   canadian-intellectual-property-statistics/ip-horizons-download-intellectual-property-data
   Download the "Researcher dataset" CSV bundle, unzip it, and point
   --patent-dir at the folder containing PT_main*.csv, PT_interested_party*.csv,
   and PT_IPC_classification*.csv.

2. CIPO IP Horizons bulk TRADEMARK data -> fully automatable (free CSV/XML download)
   Same page as above, trademark researcher dataset.
   Point --trademark-dir at the unzipped trademark folder.

3. WIPO PATENTSCOPE -> NOT freely API-accessible.
   WIPO's only programmatic interface is a paid SOAP web service (CHF 2,000/yr).
   The free path is: log into patentscope.wipo.int with a free account, run your
   search, and use the built-in "Export" feature (up to 10,000 records as CSV/XLS).
   This script reads that exported file -- it does not scrape PATENTSCOPE itself,
   since doing so programmatically would violate WIPO's authorized-use terms
   (max ~10 requests/minute per IP, no automated bulk querying).
   Point --wipo-export at the file you exported.

4. TSX / TSX Venture company list -> mostly automatable
   TMX publishes a current listed-issuers spreadsheet (with sector) at:
   https://www.tsx.com/en/listings/current-market-statistics
   The script will try to download it directly; if TMX changes the URL or
   gates it behind a click-through, download it by hand and pass it via
   --tsx-file instead.

OUTPUT
------
A single .xlsx with one row per unique company, columns:
    Company Name | Sector | Source List(s) | Agent of Record | Filing Status |
    Matched Name Variants | Hit Count
Companies found in more than one source are merged (fuzzy-matched on name)
and their sources are combined into one cell instead of duplicated rows.

"Agent of Record" / "Filing Status" flag whether a CIPO patent or trademark
filing lists a registered IP agent/firm. Filing Status is one of:
    - "Has registered agent"               -> an agent/firm is named
    - "No agent found (possible self-filer)" -> agent data was available for
      this source but no agent was listed -- the strongest BD signal
    - "Unknown (no agent data in source)"  -> agent data wasn't available to
      check (e.g. you didn't supply the agent file)
    - "N/A (no IP filing source matched)"  -> this company only showed up via
      TSX listings, which has no concept of an IP agent

USAGE
-----
    python build_prospect_list.py \
        --patent-dir path/to/cipo_patents/ \
        --trademark-dir path/to/cipo_trademarks/ \
        --wipo-export path/to/patentscope_export.csv \
        --tsx-file path/to/tsx_listed_companies.xlsx \
        --fc-dockets-csv path/to/fc_dockets.csv \
        --litigation-signals \
        --sectors ai cybersecurity cleantech \
        --output prospects.xlsx

Any source you don't have a file for yet can simply be omitted -- the
script will skip it and tell you it was skipped.

LITIGATION MODULE (optional, two layers)
-----------------------------------------
--fc-dockets-csv alone adds basic litigation-hit lookup: which prospects
are named in a Federal Court IP case, plus case number, counsel, and filing
date. On its own this just tells you who's a current litigant -- usually
already represented for that matter.

Adding --litigation-signals on top computes the more useful prospecting
signals: freshness (was this filed recently enough that counsel selection
may still be live?), counsel-specialism mismatch (represented, but not by a
firm on your known-IP-litigation-specialist list -- a referral/co-counsel
opportunity), and wave/cluster detection (a sector under active enforcement
right now -- flags companies who share the fact pattern but AREN'T yet
named, which is usually the highest-value signal of the three).

See fc_dockets_template.csv for the input file shape, and edit
KNOWN_IP_LITIGATION_FIRMS near the top of this script (or pass
--known-litigation-firms-file) to reflect who you consider a true IP
litigation specialist.
"""

import argparse
import re
import sys
from pathlib import Path

import pandas as pd
import requests
from rapidfuzz import fuzz

# --------------------------------------------------------------------------
# SECTOR DEFINITIONS
# Each sector maps to: IPC prefixes (patents), Nice classes (trademarks),
# and free-text keywords (used against titles/abstracts/descriptions and as
# a fallback for the WIPO export and TSX sector column).
#
# IPC prefixes use the 4-character format: Section (1 letter) + Class (2
# digits, zero-padded) + Subclass (1 letter). E.g. "G06N" = section G,
# class 06, subclass N (artificial intelligence). The patent loader
# reconstructs this format from CIPO's split columns and does a startswith
# match, so "G06N" catches G06N3 (neural nets), G06N20 (ML), etc.
# --------------------------------------------------------------------------
SECTOR_DEFINITIONS = {
    "ai": {
        "ipc_prefixes": ["G06N", "G06V", "G06T"],
        # G06N = computing using AI/ML, G06V = image/video understanding,
        # G06T = image processing / computer graphics
        "nice_classes": [9, 42],
        "keywords": ["artificial intelligence", "machine learning", "neural network",
                     "deep learning", "computer vision", "natural language processing",
                     " ai ", "ai-powered", "generative ai"],
        "tsx_sector_keywords": ["technology"],
    },
    "cybersecurity": {
        "ipc_prefixes": ["H04L", "H04K"],
        # H04L = data communication networks (incl. security/encryption protocols),
        # H04K = secret communication / jamming
        "nice_classes": [9, 42, 45],
        "keywords": ["cybersecurity", "cyber security", "encryption", "threat detection",
                     "network security", "data security", "infosec"],
        "tsx_sector_keywords": ["technology"],
    },
    "cleantech": {
        "ipc_prefixes": ["Y02E", "H01M", "F03D", "C25B"],
        # Y02E = energy reduction tech, H01M = electrochemical processes (batteries),
        # F03D = wind motors, C25B = electrolytic production (hydrogen)
        "nice_classes": [4, 7, 9, 37, 40, 42],
        "keywords": ["clean energy", "renewable", "battery", "hydrogen", "carbon capture",
                     "solar", "wind power", "energy storage", "decarbonization"],
        "tsx_sector_keywords": ["clean technology & renewable energy", "oil & gas", "utilities & pipelines"],
    },
    "biotech": {
        "ipc_prefixes": ["C12N", "C12Q", "A61K", "A61P"],
        # C12N = microorganisms / genetic engineering, C12Q = enzymatic/microbial assays,
        # A61K = preparations for medical use, A61P = therapeutic activity of compounds
        "nice_classes": [1, 5, 42],
        "keywords": ["biotechnology", "therapeutics", "genomics", "pharmaceutical",
                     "diagnostics", "gene therapy", "vaccine"],
        "tsx_sector_keywords": ["life sciences"],
    },
}

# --------------------------------------------------------------------------
# COLUMN NAME GUESSES
# CIPO's bulk CSVs and WIPO/TSX exports don't have perfectly stable column
# names across releases. These lists give the loader several aliases to try
# per logical field; first match wins. If a load fails to find your data,
# add your file's actual column name to the relevant list.
# --------------------------------------------------------------------------
COLUMN_ALIASES = {
    "name": ["applicant_name", "applicant", "owner_name", "owner", "party_name",
             "current_owner_name", "name", "Company", "Name", "Issuer", "Issuer Name"],
    "classification": ["ipc", "ipc_code", "classification", "nice_class",
                        "nice_classification", "cipo_classification", "class"],
    "text": ["title", "abstract", "mark_description", "description", "invention_title"],
    "sector": ["Sector", "sector", "Industry", "industry"],
    "application_number": ["application_number", "app_no", "application_no", "appl_no",
                            "case_number", "filing_number", "application_num"],
    "agent_name": ["agent_name", "agent", "representative_name", "representative",
                   "party_representation", "interested_party_representation",
                   "registered_agent", "agent_of_record"],
}

# --------------------------------------------------------------------------
# LITIGATION MODULE CONFIG
#
# KNOWN_IP_LITIGATION_FIRMS: used by the optional --litigation-signals
# counsel-mismatch flag to tell "represented by a recognized IP litigation
# specialist" apart from "represented by someone else." This is a judgment
# call about reputation/specialization that you (the litigator) are far
# better positioned to make than this script -- the names below are just
# illustrative placeholders. Replace/expand this list, or maintain it
# externally and pass it via --known-litigation-firms-file (one name per
# line, merged with this list).
# --------------------------------------------------------------------------
KNOWN_IP_LITIGATION_FIRMS = [
    "Smart & Biggar",
    "Bereskin & Parr",
    "Gowling WLG",
    "Norton Rose Fulbright",
    "Belmore Neidrauer",
    "Aitken Klee",
    "Osler",
    "McCarthy Tetrault",
]

# Column aliases for the Federal Court docket import file (see
# load_fc_dockets / fc_dockets_template.csv). Kept separate from
# COLUMN_ALIASES since this is a different kind of source file.
LITIGATION_COLUMN_ALIASES = {
    "court_file_number": ["court_file_number", "court_file_no", "file_number",
                          "docket_number", "case_number"],
    "party_names": ["party_names", "parties", "party_name", "style_of_cause"],
    "counsel_names": ["counsel_names", "counsel", "legal_counsel", "lawyer", "solicitor"],
    "filing_date": ["filing_date", "filed_date", "date_filed", "filing_dt"],
    "case_reference_text": ["case_reference_text", "ip_reference", "case_title",
                            "subject_matter", "notes", "ip_number_or_name"],
    "self_represented": ["self_represented", "self_rep", "unrepresented"],
}


def find_column(df: pd.DataFrame, field: str, aliases: dict | None = None) -> str | None:
    """Return the first matching column name in df for a logical field."""
    aliases = aliases if aliases is not None else COLUMN_ALIASES
    cols_lower = {c.lower(): c for c in df.columns}
    for alias in aliases[field]:
        if alias.lower() in cols_lower:
            return cols_lower[alias.lower()]
    return None


def _find_col_substr(df: pd.DataFrame, *substrings: str) -> str | None:
    """Return first column whose name contains any of the given substrings (case-insensitive)."""
    for col in df.columns:
        col_lower = col.lower()
        if any(s.lower() in col_lower for s in substrings):
            return col
    return None


def _read_pipe_csv(path: Path) -> pd.DataFrame | None:
    """Read a CIPO pipe-delimited CSV trying multiple encodings."""
    for enc in ("utf-8", "latin-1", "cp1252", "utf-16", "utf-16-le"):
        try:
            df = pd.read_csv(path, sep="|", encoding=enc, low_memory=False, on_bad_lines="skip")
            if len(df.columns) > 1:
                return df
        except Exception:
            continue
    return None


def normalize_name(name: str) -> str:
    """Lowercase, strip legal suffixes and punctuation for fuzzy matching."""
    if not isinstance(name, str):
        return ""
    s = name.lower()
    s = re.sub(r"[.,]", "", s)
    suffixes = [r"\binc\b", r"\bincorporated\b", r"\bltd\b", r"\blimited\b",
                r"\bcorp\b", r"\bcorporation\b", r"\bllc\b", r"\bllp\b",
                r"\bco\b", r"\bcompany\b", r"\bplc\b", r"\bs\.a\.\b"]
    for suf in suffixes:
        s = re.sub(suf, "", s)
    s = re.sub(r"[^a-z0-9 ]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def matches_sector(row_text: str, sector_keywords: list[str]) -> bool:
    """Substring match -- for free text fields like titles/abstracts/descriptions."""
    row_text = (row_text or "").lower()
    return any(kw.lower() in row_text for kw in sector_keywords)


def matches_tsx_category(sector_val: str, sector_keywords: list[str]) -> bool:
    """
    Exact match -- TSX's Sector column is a fixed category (e.g. "Technology",
    "Clean Technology & Renewable Energy"), so substring matching would wrongly
    match "Technology" inside "Clean Technology & Renewable Energy". Compare
    the whole trimmed/lowercased category string instead.
    """
    sector_val = (sector_val or "").strip().lower()
    return any(sector_val == kw.lower() for kw in sector_keywords)


# --------------------------------------------------------------------------
# SOURCE LOADERS
# Each returns a list of dicts: {"name": str, "sector": str, "source": str}
# --------------------------------------------------------------------------

def load_cipo_patents(patent_dir: str, sectors: list[str]) -> list[dict]:
    """
    Loads CIPO IP Horizons bulk patent data from a directory containing:
      PT_main*.csv               – patent metadata (title, status, filing date)
      PT_interested_party*.csv   – owner / inventor / agent names (pipe-delimited)
      PT_IPC_classification*.csv – IPC classification codes (pipe-delimited)

    All files are pipe-delimited with bilingual English-French column headers.
    Owners are joined from PT_interested_party where Interested Party Type Code = OWNR.
    Agents are joined where type = AGNT.
    IPC prefix is reconstructed as Section + zero-padded Class + Subclass (e.g. "G06N").
    """
    print(f"[CIPO Patents] Loading from: {patent_dir}")
    d = Path(patent_dir)
    if not d.is_dir():
        print(f"[CIPO Patents] '{patent_dir}' is not a directory. "
              "Pass the folder containing PT_main*.csv, PT_interested_party*.csv, "
              "and PT_IPC_classification*.csv.")
        return []

    def find_file(*patterns):
        for pat in patterns:
            matches = sorted(d.glob(pat))
            if matches:
                return matches[0]
        return None

    main_path  = find_file("PT_main*.csv",  "pt_main*.csv")
    party_path = find_file("PT_interested_party*.csv", "pt_interested_party*.csv")
    ipc_path   = find_file("PT_IPC_classification*.csv", "pt_ipc_classification*.csv", "pt_ipc*.csv")

    if not main_path:
        print("[CIPO Patents] Could not find PT_main*.csv in the given directory.")
        return []

    # ── Read PT_main ──────────────────────────────────────────────────────────
    print(f"[CIPO Patents]   Reading {main_path.name} ...")
    main_df = _read_pipe_csv(main_path)
    if main_df is None or main_df.empty:
        print("[CIPO Patents] Failed to read PT_main file.")
        return []

    pn_col    = _find_col_substr(main_df, "patent number")
    title_col = _find_col_substr(main_df, "title english", "english title")

    if not pn_col:
        print(f"[CIPO Patents] Could not find patent-number column. "
              f"Columns seen: {list(main_df.columns[:8])}")
        return []

    print(f"[CIPO Patents]   {len(main_df):,} patents in PT_main.")

    # ── Read PT_interested_party ──────────────────────────────────────────────
    # OWNR rows give us the applicant/owner company name.
    # AGNT rows give us the IP agent of record.
    owner_lookup: dict[str, str] = {}
    agent_lookup: dict[str, str] = {}
    agent_data_available = False

    if party_path:
        print(f"[CIPO Patents]   Reading {party_path.name} ...")
        party_df = _read_pipe_csv(party_path)
        if party_df is not None:
            p_pn_col   = _find_col_substr(party_df, "patent number")
            p_name_col = _find_col_substr(party_df, "party name")
            p_type_col = _find_col_substr(party_df, "interested party type code", "type code")

            if p_pn_col and p_name_col and p_type_col:
                agent_data_available = True
                for _, row in party_df.iterrows():
                    ptype = str(row.get(p_type_col, "")).strip().upper()
                    pname = str(row.get(p_name_col, "")).strip()
                    pn    = str(row.get(p_pn_col, "")).strip()
                    if not pn or not pname or pname in ("nan", "None", ""):
                        continue
                    if ptype == "OWNR":
                        owner_lookup.setdefault(pn, pname)
                    elif ptype == "AGNT":
                        agent_lookup.setdefault(pn, pname)
                print(f"[CIPO Patents]   {len(owner_lookup):,} owners, "
                      f"{len(agent_lookup):,} agents loaded from PT_interested_party.")
            else:
                print(f"[CIPO Patents]   PT_interested_party: could not locate required columns. "
                      f"Columns seen: {list(party_df.columns[:8])}")
    else:
        print("[CIPO Patents]   PT_interested_party not found -- owner/agent data unavailable.")

    # ── Read PT_IPC_classification ────────────────────────────────────────────
    # Reconstruct 4-char IPC prefix: Section + zero-padded Class + Subclass
    # e.g. section="G", class="6", subclass="N" -> "G06N"
    pn_to_ipc4: dict[str, set[str]] = {}  # patent_number -> set of 4-char IPC prefixes

    if ipc_path:
        print(f"[CIPO Patents]   Reading {ipc_path.name} ...")
        ipc_df = _read_pipe_csv(ipc_path)
        if ipc_df is not None:
            i_pn_col  = _find_col_substr(ipc_df, "patent number")
            i_sec_col = _find_col_substr(ipc_df, "ipc section code")
            i_cls_col = _find_col_substr(ipc_df, "ipc class code")
            i_sub_col = _find_col_substr(ipc_df, "ipc subclass code")

            if i_pn_col and i_sec_col and i_cls_col and i_sub_col:
                for _, row in ipc_df.iterrows():
                    pn  = str(row.get(i_pn_col,  "")).strip()
                    sec = str(row.get(i_sec_col, "")).strip()
                    cls = str(row.get(i_cls_col, "")).strip()
                    sub = str(row.get(i_sub_col, "")).strip()
                    if not (pn and sec and cls and sub):
                        continue
                    try:
                        ipc4 = f"{sec}{int(cls):02d}{sub}".upper()
                    except ValueError:
                        continue
                    pn_to_ipc4.setdefault(pn, set()).add(ipc4)
                print(f"[CIPO Patents]   {len(pn_to_ipc4):,} patents with IPC codes loaded.")
            else:
                print(f"[CIPO Patents]   PT_IPC_classification: could not locate IPC columns. "
                      f"Columns seen: {list(ipc_df.columns[:8])}")
    else:
        print("[CIPO Patents]   PT_IPC_classification not found -- IPC sector filtering unavailable.")

    # ── Match sectors per patent ──────────────────────────────────────────────
    pn_to_sectors: dict[str, set[str]] = {}
    for sector in sectors:
        sdef = SECTOR_DEFINITIONS[sector]
        prefixes = [p.upper() for p in sdef["ipc_prefixes"]]
        for pn, ipc4_set in pn_to_ipc4.items():
            for ipc4 in ipc4_set:
                if any(ipc4.startswith(p) for p in prefixes):
                    pn_to_sectors.setdefault(pn, set()).add(sector)
                    break

    # ── Build output records ──────────────────────────────────────────────────
    records = []
    for _, row in main_df.iterrows():
        pn = str(row.get(pn_col, "")).strip()
        if not pn:
            continue

        matched_sectors = set(pn_to_sectors.get(pn, set()))

        # Keyword fallback against English title when IPC gives no match
        if title_col:
            title_val = str(row.get(title_col, ""))
            for sector in sectors:
                if sector not in matched_sectors:
                    if matches_sector(title_val, SECTOR_DEFINITIONS[sector]["keywords"]):
                        matched_sectors.add(sector)

        if not matched_sectors:
            continue

        owner_name = owner_lookup.get(pn)
        agent_name = agent_lookup.get(pn)

        if not owner_name:
            continue  # no owner record means we have nothing to prospect

        if agent_name:
            agent_status = "represented"
        elif agent_data_available:
            agent_status = "self"
        else:
            agent_status = "unknown"

        for sector in sorted(matched_sectors):
            records.append({
                "name": owner_name,
                "sector": sector,
                "source": "CIPO Patents (IP Horizons)",
                "agent": agent_name,
                "agent_status": agent_status,
            })

    print(f"[CIPO Patents] {len(records)} sector-matching records found.")
    return records


def load_cipo_trademarks(trademark_dir: str, sectors: list[str]) -> list[dict]:
    """
    Loads CIPO IP Horizons bulk trademark data from a directory containing:
      TM_application_main*.csv    – application metadata and Nice class
      TM_interested_party*.csv    – applicant / owner / agent names (pipe-delimited, UTF-16)

    All files are pipe-delimited with bilingual English-French column headers.
    """
    print(f"[CIPO Trademarks] Loading from: {trademark_dir}")
    d = Path(trademark_dir)
    if not d.is_dir():
        print(f"[CIPO Trademarks] '{trademark_dir}' is not a directory. "
              "Pass the folder containing TM_application_main*.csv and TM_interested_party*.csv.")
        return []

    def find_file(*patterns):
        for pat in patterns:
            matches = sorted(d.glob(pat))
            if matches:
                return matches[0]
        return None

    main_path  = find_file("TM_application_main*.csv", "tm_application_main*.csv")
    party_path = find_file("TM_interested_party*.csv", "tm_interested_party*.csv")

    if not main_path:
        print("[CIPO Trademarks] Could not find TM_application_main*.csv in the given directory.")
        return []

    # ── Read TM_application_main ──────────────────────────────────────────────
    print(f"[CIPO Trademarks]   Reading {main_path.name} ...")
    main_df = _read_pipe_csv(main_path)
    if main_df is None or main_df.empty:
        print("[CIPO Trademarks] Failed to read TM_application_main file.")
        return []

    app_col   = _find_col_substr(main_df, "application number")
    nice_col  = _find_col_substr(main_df, "nice classification code", "nice class")
    text_col  = _find_col_substr(main_df, "mark description", "description", "title")

    if not app_col:
        print(f"[CIPO Trademarks] Could not find application-number column. "
              f"Columns seen: {list(main_df.columns[:8])}")
        return []

    print(f"[CIPO Trademarks]   {len(main_df):,} applications in TM_application_main.")

    # ── Read TM_interested_party ──────────────────────────────────────────────
    # Party Type Code 10 = current owner/registrant (the prospecting target).
    # Current Owner Legal Name column is also available and preferred where not "Unknown".
    owner_lookup: dict[str, str] = {}
    agent_lookup: dict[str, str] = {}
    agent_data_available = False

    if party_path:
        print(f"[CIPO Trademarks]   Reading {party_path.name} ...")
        party_df = _read_pipe_csv(party_path)
        if party_df is not None:
            p_app_col   = _find_col_substr(party_df, "application number")
            p_name_col  = _find_col_substr(party_df, "party name")
            p_owner_col = _find_col_substr(party_df, "current owner legal name")
            p_type_col  = _find_col_substr(party_df, "party type code")
            p_agent_col = _find_col_substr(party_df, "agent number")

            if p_app_col and (p_name_col or p_owner_col):
                agent_data_available = True
                for _, row in party_df.iterrows():
                    app_no = str(row.get(p_app_col, "")).strip()
                    if not app_no:
                        continue
                    ptype = str(row.get(p_type_col, "")).strip() if p_type_col else ""

                    # Prefer Current Owner Legal Name; fall back to Party Name
                    owner_name = None
                    if p_owner_col:
                        val = str(row.get(p_owner_col, "")).strip()
                        if val and val not in ("nan", "None", "Unknown", "-1"):
                            owner_name = val
                    if not owner_name and p_name_col:
                        val = str(row.get(p_name_col, "")).strip()
                        if val and val not in ("nan", "None", ""):
                            owner_name = val

                    if owner_name:
                        owner_lookup.setdefault(app_no, owner_name)

                    # Agent number present and valid means represented
                    if p_agent_col:
                        agent_val = str(row.get(p_agent_col, "")).strip()
                        if agent_val and agent_val not in ("-1", "nan", "None", ""):
                            agent_lookup.setdefault(app_no, f"Agent #{agent_val}")

                print(f"[CIPO Trademarks]   {len(owner_lookup):,} owners, "
                      f"{len(agent_lookup):,} agent records loaded.")
            else:
                print(f"[CIPO Trademarks]   TM_interested_party: could not locate required columns. "
                      f"Columns seen: {list(party_df.columns[:8])}")
    else:
        print("[CIPO Trademarks]   TM_interested_party not found -- owner/agent data unavailable.")

    # ── Build output records ──────────────────────────────────────────────────
    records = []
    for _, row in main_df.iterrows():
        app_no = str(row.get(app_col, "")).strip()
        if not app_no:
            continue

        owner_name = owner_lookup.get(app_no)
        if not owner_name:
            continue

        nice_val = str(row.get(nice_col, "")) if nice_col else ""
        text_val = str(row.get(text_col, "")) if text_col else ""
        agent_name = agent_lookup.get(app_no)

        if agent_name:
            agent_status = "represented"
        elif agent_data_available:
            agent_status = "self"
        else:
            agent_status = "unknown"

        for sector in sectors:
            sdef = SECTOR_DEFINITIONS[sector]
            nice_hit = False
            if nice_val and nice_val not in ("-1", "nan", "None"):
                for nc in sdef["nice_classes"]:
                    if str(nc) in re.findall(r"\d+", nice_val):
                        nice_hit = True
                        break
            kw_hit = matches_sector(text_val, sdef["keywords"])
            if nice_hit or kw_hit:
                records.append({
                    "name": owner_name,
                    "sector": sector,
                    "source": "CIPO Trademarks (IP Horizons)",
                    "agent": agent_name,
                    "agent_status": agent_status,
                })

    print(f"[CIPO Trademarks] {len(records)} sector-matching records found.")
    return records


def load_wipo_export(export_path: str, sectors: list[str]) -> list[dict]:
    """
    Reads a manually-exported PATENTSCOPE search result file (CSV or XLS/XLSX).
    Filters to applicants that look Canadian where a country field is present,
    and tags sector by keyword match against title/abstract.
    """
    print(f"[WIPO PATENTSCOPE export] Loading {export_path} ...")
    path = Path(export_path)
    try:
        if path.suffix.lower() in (".xlsx", ".xls"):
            df = pd.read_excel(export_path)
        else:
            df = pd.read_csv(export_path, low_memory=False, on_bad_lines="skip")
    except Exception as e:
        print(f"[WIPO PATENTSCOPE export] FAILED to load: {e}")
        return []

    name_col = find_column(df, "name")
    text_col = find_column(df, "text")
    agent_col = find_column(df, "agent_name")
    country_col = None
    for c in df.columns:
        if "country" in c.lower():
            country_col = c
            break

    if not name_col:
        print(f"[WIPO PATENTSCOPE export] Could not find an applicant column among: {list(df.columns)}")
        return []
    if not agent_col:
        print("[WIPO PATENTSCOPE export] No agent/representative column found in this export -- "
              "agent-of-record will be reported as Unknown for this source.")

    records = []
    for _, row in df.iterrows():
        name = row.get(name_col)
        if not isinstance(name, str) or not name.strip():
            continue
        if country_col is not None:
            country_val = str(row.get(country_col, "")).upper()
            if country_val and "CA" not in country_val and "CANADA" not in country_val:
                continue
        text_val = str(row.get(text_col, "")) if text_col else ""

        agent_name = None
        if agent_col:
            val = row.get(agent_col)
            agent_name = val.strip() if isinstance(val, str) and val.strip() else None
        agent_status = "represented" if agent_name else ("self" if agent_col else "unknown")

        for sector in sectors:
            sdef = SECTOR_DEFINITIONS[sector]
            if matches_sector(text_val, sdef["keywords"]):
                records.append({"name": name.strip(), "sector": sector,
                                 "source": "WIPO PATENTSCOPE (manual export)",
                                 "agent": agent_name, "agent_status": agent_status})
    print(f"[WIPO PATENTSCOPE export] {len(records)} sector-matching records found.")
    return records


def fetch_tsx_listed_companies(tsx_file: str | None, sectors: list[str]) -> list[dict]:
    """
    Loads the TSX/TSXV current listed-issuers file.

    The TMX Excel file has a multi-row disclaimer block at the top before the
    real column headers. The loader scans the first 10 rows to find the header
    row (the one that contains "Co_ID" or similar identifier-like values), then
    re-reads the file with the correct header row index.
    """
    TMX_STATS_PAGE = "https://www.tsx.com/en/listings/current-market-statistics"

    if not tsx_file:
        print("[TSX Listings] No --tsx-file given, attempting direct download...")
        try:
            resp = requests.get(TMX_STATS_PAGE, timeout=15)
            resp.raise_for_status()
            print("[TSX Listings] Could not auto-locate the download link from the page "
                  "(TMX serves this as a click-through download, not a stable direct URL).")
        except Exception as e:
            print(f"[TSX Listings] Direct fetch failed: {e}")
        print(f"[TSX Listings] Please download the spreadsheet manually from:\n  {TMX_STATS_PAGE}")
        print("  then rerun with --tsx-file pointing at the downloaded .xlsx")
        return []

    print(f"[TSX Listings] Loading {tsx_file} ...")
    tsx_path = str(tsx_file)
    try:
        if tsx_path.lower().endswith(".txt"):
            df = pd.read_csv(tsx_path, sep=None, engine="python", on_bad_lines="skip")
        else:
            # The TMX xlsx has 3-4 disclaimer rows before the real header.
            # Scan the first 10 rows to find the one that looks like a header
            # (contains a short identifier-like value in the first non-null cell).
            probe = pd.read_excel(tsx_path, header=None, nrows=10)
            header_row = 3  # safe default
            for i, row in probe.iterrows():
                first_val = str(row.dropna().iloc[0]).strip() if not row.dropna().empty else ""
                # The real header row has short token-like values (Co_ID, Exchange, etc.)
                if first_val and len(first_val) <= 20 and not first_val.startswith("This") and not first_val.startswith("#"):
                    header_row = int(i)
                    break
            df = pd.read_excel(tsx_path, header=header_row)
    except Exception as e:
        print(f"[TSX Listings] FAILED to load: {e}")
        return []

    # The real column names are in the header row; find name and sector columns
    # by substring matching since TMX's exact column names shift between releases.
    name_col   = _find_col_substr(df, "name", "issuer") or find_column(df, "name")
    sector_col = _find_col_substr(df, "sector") or find_column(df, "sector")

    if not name_col:
        print(f"[TSX Listings] Could not find a company-name column among: {list(df.columns[:15])}")
        return []

    records = []
    for _, row in df.iterrows():
        name = row.get(name_col)
        if not isinstance(name, str) or not name.strip():
            continue
        sector_val = str(row.get(sector_col, "")) if sector_col else ""

        for sector in sectors:
            sdef = SECTOR_DEFINITIONS[sector]
            if matches_tsx_category(sector_val, sdef["tsx_sector_keywords"]):
                records.append({"name": name.strip(), "sector": sector,
                                 "source": "TSX/TSXV Listings",
                                 "agent": None, "agent_status": "n/a"})
    print(f"[TSX Listings] {len(records)} sector-matching records found.")
    return records


# --------------------------------------------------------------------------
# CONSOLIDATION / DEDUPE
# --------------------------------------------------------------------------

def consolidate(records: list[dict], fuzzy_threshold: int = 90) -> pd.DataFrame:
    """
    Groups records into unique companies. Exact matches are grouped on the
    normalized name; near-duplicates (e.g. "ACME Inc" vs "Acme Incorporated")
    are merged if their fuzzy match score exceeds fuzzy_threshold.
    Sources and sectors found for a company are combined into one row.
    """
    if not records:
        return pd.DataFrame(columns=["Company Name", "Sector", "Source List(s)",
                                      "Agent of Record", "Filing Status",
                                      "Matched Name Variants", "Hit Count"])

    df = pd.DataFrame(records)
    df["norm_name"] = df["name"].apply(normalize_name)
    df = df[df["norm_name"] != ""]

    groups = []          # list of dicts: {names:set, norm_names:set, sectors:set, sources:set, count:int}
    norm_to_group = {}   # exact norm_name -> group index

    unique_norms = df["norm_name"].drop_duplicates().tolist()

    for norm in unique_norms:
        if norm in norm_to_group:
            continue
        best_group_idx = None
        best_score = 0
        for idx, g in enumerate(groups):
            for existing_norm in g["norm_names"]:
                score = fuzz.token_sort_ratio(norm, existing_norm)
                if score > best_score:
                    best_score = score
                    best_group_idx = idx
        if best_score >= fuzzy_threshold:
            groups[best_group_idx]["norm_names"].add(norm)
            norm_to_group[norm] = best_group_idx
        else:
            groups.append({"norm_names": {norm}})
            norm_to_group[norm] = len(groups) - 1

    rows = []
    for idx in range(len(groups)):
        norms_in_group = groups[idx]["norm_names"]
        sub = df[df["norm_name"].isin(norms_in_group)]
        display_name = sub["name"].value_counts().idxmax()
        name_variants = sorted(set(sub["name"]) - {display_name})
        sectors_found = sorted(set(sub["sector"]))
        sources_found = sorted(set(sub["source"]))

        agents_found = sorted({a for a in sub["agent"] if isinstance(a, str) and a.strip()})
        statuses = set(sub["agent_status"]) if "agent_status" in sub.columns else set()
        if agents_found:
            filing_status = "Has registered agent"
        elif "self" in statuses:
            filing_status = "No agent found (possible self-filer)"
        elif statuses and statuses.issubset({"n/a"}):
            filing_status = "N/A (no IP filing source matched)"
        else:
            filing_status = "Unknown (no agent data in source)"

        rows.append({
            "Company Name": display_name,
            "Sector": ", ".join(sectors_found),
            "Source List(s)": ", ".join(sources_found),
            "Agent of Record": "; ".join(agents_found),
            "Filing Status": filing_status,
            "Matched Name Variants": "; ".join(name_variants) if name_variants else "",
            "Hit Count": len(sub),
        })

    result = pd.DataFrame(rows).sort_values(
        by=["Hit Count", "Company Name"], ascending=[False, True]
    ).reset_index(drop=True)
    return result


def write_output(df: pd.DataFrame, output_path: str) -> None:
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Prospects")
        ws = writer.sheets["Prospects"]

        header_font = Font(bold=True, color="FFFFFF", name="Arial")
        header_fill = PatternFill("solid", start_color="2F5597")
        for col_idx, _ in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        widths = {"Company Name": 38, "Sector": 18, "Source List(s)": 32,
                  "Agent of Record": 30, "Filing Status": 30,
                  "Matched Name Variants": 40, "Hit Count": 10,
                  "Litigation Hit": 14, "FC Case Number(s)": 22,
                  "Litigation Counsel": 30, "Litigation Filing Date(s)": 24,
                  "Litigation Freshness": 38, "Counsel Specialism Flag": 38,
                  "Litigation Cluster Flag": 50}
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 20)

        ws.freeze_panes = "A2"

    print(f"\nSaved {len(df)} consolidated companies to {output_path}")


# --------------------------------------------------------------------------
# LITIGATION MODULE
# --------------------------------------------------------------------------

def load_fc_dockets(csv_path: str) -> pd.DataFrame:
    print(f"[FC Dockets] Loading {csv_path} ...")
    try:
        raw = pd.read_csv(csv_path, low_memory=False, on_bad_lines="skip")
    except Exception as e:
        print(f"[FC Dockets] FAILED to load: {e}")
        return pd.DataFrame()

    col_map = {}
    for field in LITIGATION_COLUMN_ALIASES:
        col = find_column(raw, field, aliases=LITIGATION_COLUMN_ALIASES)
        if col:
            col_map[field] = col

    if "party_names" not in col_map or "court_file_number" not in col_map:
        print(f"[FC Dockets] Could not find party-names / court-file-number columns among: {list(raw.columns)}")
        print("  -> Add the real column names to LITIGATION_COLUMN_ALIASES and rerun.")
        return pd.DataFrame()

    df = raw.rename(columns={v: k for k, v in col_map.items()})
    if "filing_date" in df.columns:
        df["filing_date"] = pd.to_datetime(df["filing_date"], errors="coerce")
    print(f"[FC Dockets] {len(df)} case record(s) loaded.")
    return df


def split_parties(party_cell) -> list[str]:
    """Splits a 'Acme Inc. v. Beta Corp.' / 'Acme Inc.; Gamma Ltd.' style cell into names."""
    if not isinstance(party_cell, str) or not party_cell.strip():
        return []
    sides = re.split(r"\s+v\.?\s+|\s+vs\.?\s+", party_cell, flags=re.IGNORECASE)
    names = []
    for side in sides:
        names.extend([p.strip() for p in re.split(r"[;,]", side) if p.strip()])
    return names


def build_party_index(fc_df: pd.DataFrame) -> dict:
    """normalized party name -> list of fc_df row indices that name them."""
    index = {}
    if fc_df.empty:
        return index
    for idx, row in fc_df.iterrows():
        for party in split_parties(row.get("party_names")):
            norm = normalize_name(party)
            if norm:
                index.setdefault(norm, []).append(idx)
    return index


def find_litigation_matches(company_name_variants: list[str], party_index: dict, fuzzy_threshold: int) -> set:
    matched = set()
    norm_variants = [normalize_name(n) for n in company_name_variants if n]
    for norm_party, idxs in party_index.items():
        for norm_variant in norm_variants:
            if norm_variant and fuzz.token_sort_ratio(norm_variant, norm_party) >= fuzzy_threshold:
                matched.update(idxs)
                break
    return matched


def apply_litigation_lookup(prospect_df: pd.DataFrame, fc_df: pd.DataFrame, fuzzy_threshold: int) -> pd.DataFrame:
    prospect_df = prospect_df.copy()
    prospect_df["Litigation Hit"] = "No"
    prospect_df["FC Case Number(s)"] = ""
    prospect_df["Litigation Counsel"] = ""
    prospect_df["Litigation Filing Date(s)"] = ""
    prospect_df["_case_idxs"] = [set() for _ in range(len(prospect_df))]

    if fc_df.empty:
        print("[FC Dockets] No usable case data -- litigation columns left blank.")
        return prospect_df

    party_index = build_party_index(fc_df)

    for i, row in prospect_df.iterrows():
        variants = [row["Company Name"]] + [v for v in str(row.get("Matched Name Variants", "")).split("; ") if v]
        case_idxs = find_litigation_matches(variants, party_index, fuzzy_threshold)
        if not case_idxs:
            continue
        cases = fc_df.loc[sorted(case_idxs)]
        prospect_df.at[i, "_case_idxs"] = case_idxs
        prospect_df.at[i, "Litigation Hit"] = "Yes"
        prospect_df.at[i, "FC Case Number(s)"] = "; ".join(sorted(set(cases["court_file_number"].astype(str))))
        if "counsel_names" in cases.columns:
            counsel_vals = sorted(set(cases["counsel_names"].dropna().astype(str)))
            prospect_df.at[i, "Litigation Counsel"] = "; ".join(counsel_vals)
        if "filing_date" in cases.columns:
            dates = sorted(set(cases["filing_date"].dropna().dt.strftime("%Y-%m-%d")))
            prospect_df.at[i, "Litigation Filing Date(s)"] = "; ".join(dates)

    n_hits = (prospect_df["Litigation Hit"] == "Yes").sum()
    print(f"[FC Dockets] {n_hits} companies matched to at least one Federal Court IP case.")
    return prospect_df


def apply_litigation_signals(prospect_df: pd.DataFrame, fc_df: pd.DataFrame,
                              known_firms: list[str], freshness_days: int,
                              cluster_window_days: int, cluster_min_size: int) -> pd.DataFrame:
    prospect_df = prospect_df.copy()
    prospect_df["Litigation Freshness"] = ""
    prospect_df["Counsel Specialism Flag"] = ""
    prospect_df["Litigation Cluster Flag"] = ""

    if fc_df.empty:
        print("[Litigation Signals] No usable case data -- signal columns left blank.")
        prospect_df.drop(columns=["_case_idxs"], inplace=True, errors="ignore")
        return prospect_df

    today = pd.Timestamp.now().normalize()
    known_firms_lower = [f.lower() for f in known_firms]
    has_self_rep_col = "self_represented" in fc_df.columns

    for i, row in prospect_df.iterrows():
        case_idxs = row["_case_idxs"]
        if not case_idxs:
            continue
        cases = fc_df.loc[sorted(case_idxs)]

        if "filing_date" in cases.columns and cases["filing_date"].notna().any():
            most_recent = cases["filing_date"].max()
            age_days = (today - most_recent).days
            if age_days <= freshness_days:
                prospect_df.at[i, "Litigation Freshness"] = (
                    f"Filed {age_days}d ago (<= {freshness_days}d) -- counsel selection may still be in progress")
            else:
                prospect_df.at[i, "Litigation Freshness"] = f"Established case (filed {age_days}d ago)"

        self_rep_hit = False
        if has_self_rep_col:
            vals = cases["self_represented"].astype(str).str.lower()
            self_rep_hit = vals.isin(["true", "yes", "y", "1"]).any()

        if self_rep_hit:
            prospect_df.at[i, "Counsel Specialism Flag"] = "Self-represented party"
        elif "counsel_names" in cases.columns and cases["counsel_names"].notna().any():
            counsel_blob = " | ".join(cases["counsel_names"].dropna().astype(str)).lower()
            if any(firm in counsel_blob for firm in known_firms_lower):
                prospect_df.at[i, "Counsel Specialism Flag"] = "Represented by a recognized IP litigation specialist"
            else:
                prospect_df.at[i, "Counsel Specialism Flag"] = (
                    "Represented, but not by a firm on your known-specialist list -- "
                    "possible referral/co-counsel opportunity")
        else:
            prospect_df.at[i, "Counsel Specialism Flag"] = "Unknown (no counsel data in docket file)"

    prospect_df.drop(columns=["_case_idxs"], inplace=True, errors="ignore")

    clusters = detect_litigation_clusters(fc_df, cluster_window_days, cluster_min_size)
    if clusters:
        prospect_df = apply_cluster_flags(prospect_df, clusters)
    else:
        print("[Litigation Signals] No clusters met the size/window threshold "
              f"(>= {cluster_min_size} cases within {cluster_window_days} days) -- "
              "Litigation Cluster Flag left blank for all rows.")

    return prospect_df


def detect_litigation_clusters(fc_df: pd.DataFrame, window_days: int, min_size: int) -> list[dict]:
    clusters = []
    if fc_df.empty or "filing_date" not in fc_df.columns or "case_reference_text" not in fc_df.columns:
        return clusters

    text_lower = fc_df["case_reference_text"].fillna("").str.lower()
    for sector, sdef in SECTOR_DEFINITIONS.items():
        mask = text_lower.apply(lambda t: any(kw.lower() in t for kw in sdef["keywords"]))
        sector_cases = fc_df[mask].dropna(subset=["filing_date"]).sort_values("filing_date")
        if len(sector_cases) < min_size:
            continue

        best = None
        dates = sector_cases["filing_date"].tolist()
        for start_date in dates:
            window_end = start_date + pd.Timedelta(days=window_days)
            in_window = sector_cases[(sector_cases["filing_date"] >= start_date) &
                                      (sector_cases["filing_date"] <= window_end)]
            if len(in_window) >= min_size and (best is None or len(in_window) > best["case_count"]):
                named = set()
                for cell in in_window["party_names"]:
                    named.update(split_parties(cell))
                best = {
                    "sector": sector,
                    "start_date": in_window["filing_date"].min(),
                    "end_date": in_window["filing_date"].max(),
                    "case_count": len(in_window),
                    "named_parties": named,
                }
        if best:
            clusters.append(best)
            print(f"[Litigation Signals] Cluster found: {best['case_count']} {sector} case(s) "
                  f"between {best['start_date'].date()} and {best['end_date'].date()}.")
    return clusters


def apply_cluster_flags(prospect_df: pd.DataFrame, clusters: list[dict]) -> pd.DataFrame:
    for c in clusters:
        named_norm = {normalize_name(p) for p in c["named_parties"]}
        for i, row in prospect_df.iterrows():
            row_sectors = [s.strip() for s in str(row.get("Sector", "")).split(",")]
            if c["sector"] not in row_sectors:
                continue
            if row.get("Litigation Hit") == "Yes":
                continue
            variants = [row["Company Name"]] + [v for v in str(row.get("Matched Name Variants", "")).split("; ") if v]
            variant_norms = {normalize_name(v) for v in variants}
            if variant_norms & named_norm:
                continue
            note = (f"In litigation blast radius: {c['case_count']} {c['sector']} FC IP case(s) filed "
                    f"{c['start_date'].date()}-{c['end_date'].date()}, not yet named")
            existing = prospect_df.at[i, "Litigation Cluster Flag"]
            prospect_df.at[i, "Litigation Cluster Flag"] = f"{existing} | {note}" if existing else note
    return prospect_df


# --------------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Build a deduped IP-prospect list from public sources.")
    parser.add_argument("--patent-dir",
                        help="Path to the folder containing CIPO IP Horizons patent CSV files "
                             "(PT_main*.csv, PT_interested_party*.csv, PT_IPC_classification*.csv)")
    parser.add_argument("--trademark-dir",
                        help="Path to the folder containing CIPO IP Horizons trademark CSV files "
                             "(TM_application_main*.csv, TM_interested_party*.csv)")
    parser.add_argument("--wipo-export", help="Path to a manually-exported PATENTSCOPE CSV/XLSX")
    parser.add_argument("--tsx-file", help="Path to a downloaded TSX/TSXV listed-companies file (.xlsx or .txt)")
    parser.add_argument("--sectors", nargs="+", default=list(SECTOR_DEFINITIONS.keys()),
                         choices=list(SECTOR_DEFINITIONS.keys()),
                         help=f"Sectors to filter for. Choices: {list(SECTOR_DEFINITIONS.keys())}")
    parser.add_argument("--fuzzy-threshold", type=int, default=90,
                         help="Name-similarity score (0-100) above which two names are merged as one company")
    parser.add_argument("--fc-dockets-csv", help="Path to your compiled Federal Court IP docket file "
                                                   "(see fc_dockets_template.csv). Enables basic litigation-hit "
                                                   "lookup (Litigation Hit / FC Case Number(s) / Litigation Counsel "
                                                   "/ Litigation Filing Date(s)).")
    parser.add_argument("--litigation-signals", action="store_true",
                         help="Requires --fc-dockets-csv. Adds freshness, counsel-specialism-mismatch, and "
                              "wave/cluster ('blast radius') flags on top of the basic litigation-hit lookup.")
    parser.add_argument("--known-litigation-firms-file",
                         help="Optional text file, one IP litigation specialist firm name per line, merged with "
                              "the built-in KNOWN_IP_LITIGATION_FIRMS list for the counsel-mismatch flag.")
    parser.add_argument("--litigation-freshness-days", type=int, default=60,
                         help="A case filed within this many days is flagged as 'counsel selection may still be "
                              "in progress'. Default 60.")
    parser.add_argument("--litigation-cluster-window-days", type=int, default=90,
                         help="Width of the rolling window used to detect a 'wave' of same-sector filings. Default 90.")
    parser.add_argument("--litigation-cluster-min-size", type=int, default=3,
                         help="Minimum number of same-sector cases within the window to count as a cluster. Default 3.")
    parser.add_argument("--output", default="prospects.xlsx", help="Output .xlsx path")
    args = parser.parse_args()

    if not any([args.patent_dir, args.trademark_dir, args.wipo_export, args.tsx_file]):
        print("No input files given -- attempting TSX auto-download only.")
        print("Pass at least one of --patent-dir / --trademark-dir / --wipo-export / --tsx-file "
              "for a meaningful result.\n")

    all_records = []
    if args.patent_dir:
        all_records += load_cipo_patents(args.patent_dir, args.sectors)
    else:
        print("[CIPO Patents] Skipped (no --patent-dir given)")

    if args.trademark_dir:
        all_records += load_cipo_trademarks(args.trademark_dir, args.sectors)
    else:
        print("[CIPO Trademarks] Skipped (no --trademark-dir given)")

    if args.wipo_export:
        all_records += load_wipo_export(args.wipo_export, args.sectors)
    else:
        print("[WIPO PATENTSCOPE export] Skipped (no --wipo-export given)")

    all_records += fetch_tsx_listed_companies(args.tsx_file, args.sectors)

    print(f"\nTotal raw matching records across all sources: {len(all_records)}")
    result_df = consolidate(all_records, fuzzy_threshold=args.fuzzy_threshold)

    if args.litigation_signals and not args.fc_dockets_csv:
        print("[Litigation Signals] --litigation-signals was given without --fc-dockets-csv -- skipping "
              "the entire litigation module (nothing to compute signals from).")
    elif args.fc_dockets_csv:
        fc_df = load_fc_dockets(args.fc_dockets_csv)
        result_df = apply_litigation_lookup(result_df, fc_df, fuzzy_threshold=args.fuzzy_threshold)
        if args.litigation_signals:
            known_firms = list(KNOWN_IP_LITIGATION_FIRMS)
            if args.known_litigation_firms_file:
                try:
                    with open(args.known_litigation_firms_file) as f:
                        known_firms += [line.strip() for line in f if line.strip()]
                except Exception as e:
                    print(f"[Litigation Signals] Could not read --known-litigation-firms-file: {e}")
            result_df = apply_litigation_signals(
                result_df, fc_df, known_firms=known_firms,
                freshness_days=args.litigation_freshness_days,
                cluster_window_days=args.litigation_cluster_window_days,
                cluster_min_size=args.litigation_cluster_min_size,
            )
        else:
            result_df.drop(columns=["_case_idxs"], inplace=True, errors="ignore")

    write_output(result_df, args.output)


if __name__ == "__main__":
    sys.exit(main())
